import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string

def automation_excel(nombre_archivo):
    """
    Input sales_mes.xlsx
    Output report_mes.xlsx
    """
    archivo_excel = pd.read_excel(nombre_archivo) 

    tabla_pivot = pd.pivot_table(data=archivo_excel, index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    mes_extension = nombre_archivo.split('_')[1]
    tabla_pivot.to_excel(f'sales_{mes_extension}', startrow=3, sheet_name='Report')
    wb = load_workbook(f'sales_{mes_extension}')
    sheet = wb['Report']

    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row

    barchart = BarChart()

    data = Reference(sheet, min_col=min_col+1, max_col=max_col, min_row=min_fila, max_row=max_fila)
    categorias = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila)

    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categorias)

    sheet.add_chart(barchart, 'B10')
    barchart.style = 2
    barchart.title = 'Ventas'

    abecedario = list(string.ascii_uppercase)
    abecedario_excel = abecedario[0:max_col]

    for letra in abecedario_excel:
        if letra!='A':
            sheet[f'{letra}{max_fila+1}'] = f'=SUM({letra}{min_fila+1}:{letra}{max_fila})'
            sheet[f'{letra}{max_fila+1}'].style = 'Currency'

    sheet[f'{abecedario_excel[0]}{max_fila+1}'] = 'TOTAL'    

    sheet['A1'] = 'Reporte'
    mes = mes_extension.split('.')[0]
    sheet['A2'] = mes
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=12)


    wb.save(f'report_{mes_extension}')
    return

automation_excel('python-proyects\excel-reports-automation\sales_2023.xlsx')
